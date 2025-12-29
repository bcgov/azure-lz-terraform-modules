from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal

import pandas as pd
from azure.identity import DefaultAzureCredential
from azure.mgmt.costmanagement import CostManagementClient
from azure.mgmt.resource import SubscriptionClient
from dateutil.relativedelta import relativedelta


def get_subscription_costs(
    management_group_id: str,
    start_date: str,
    end_date: str,
    granularity: str = "Monthly",
):
    """
    Query Azure Cost Management for subscription costs within a management group.
    """
    try:
        # Initialize credentials and client
        credential = DefaultAzureCredential()
        cost_client = CostManagementClient(credential)
        sub_client = SubscriptionClient(credential)

        # Build query definition for costs
        query = {
            "type": "ActualCost",
            "timeframe": "Custom",
            "timePeriod": {
                "from": f"{start_date}T00:00:00+00:00",
                "to": f"{end_date}T23:59:59+00:00",
            },
            "dataset": {
                "granularity": granularity,
                "aggregation": {"totalCost": {"name": "Cost", "function": "Sum"}},
                "grouping": [
                    {"type": "Dimension", "name": "SubscriptionId"},
                    {"type": "Dimension", "name": "SubscriptionName"},
                ],
            },
        }

        # Execute query with management group scope
        scope = (
            f"/providers/Microsoft.Management/managementGroups/{management_group_id}"
        )
        print("\nExecuting query with parameters:")
        print(f"Scope: {scope}")
        print(f"Query parameters: {query}")

        try:
            results = cost_client.query.usage(scope=scope, parameters=query)
            print("\nQuery executed successfully")
        except Exception as api_error:
            print("\nAPI Error Details:")
            print(f"Error type: {type(api_error).__name__}")
            if hasattr(api_error, "response"):
                print(f"Response status: {api_error.response.status_code}")
                print(f"Response headers: {api_error.response.headers}")
                print(f"Response content: {api_error.response.text}")
            raise

        # Process results into DataFrame
        cost_data = []
        for row in results.rows:
            # Print row data for debugging
            print(f"Debug - Row data: {row}")
            sub_name = row[3].split("(")[0].strip()
            license_plate = sub_name.split("-")[0].strip() if "-" in sub_name else sub_name
            cost_data.append(
                {
                    "Cost": round(
                        float(row[0]), 2
                    ),  # Cost is first, rounded to 2 decimals
                    "Date": row[1],  # Date is second
                    "SubscriptionId": row[2],  # ID is third
                    "SubscriptionName": sub_name,  # Name is fourth
                    "LicensePlate": license_plate,
                    "Currency": row[4],  # Currency is last
                }
            )

        df = pd.DataFrame(cost_data)

        # Get subscription tags
        print("\nFetching subscription tags...")
        for index, row in df.iterrows():
            try:
                sub = sub_client.subscriptions.get(row["SubscriptionId"])
                df.at[index, "AccountCoding"] = (
                    sub.tags.get("account_coding", "Untagged")
                    if sub.tags
                    else "Untagged"
                )
                df.at[index, "ExpenseAuthority"] = (
                    sub.tags.get("expense_authority", "Untagged")
                    if sub.tags
                    else "Untagged"
                )
            except Exception as e:
                print(
                    f"Warning: Could not fetch tags for subscription {row['SubscriptionId']}: {str(e)}"
                )
                df.at[index, "AccountCoding"] = "Untagged"
                df.at[index, "ExpenseAuthority"] = "Untagged"

        # Create summary by account coding with tax calculations
        summary_df = (
            df.groupby(["AccountCoding", "ExpenseAuthority"])
            .agg(
                {
                    "Cost": lambda x: round(sum(x), 2),
                    "SubscriptionName": lambda x: ", ".join(sorted(set(x))),
                    "LicensePlate": lambda x: ", ".join(sorted(set(x))),
                }
            )
            .reset_index()
        )

        # Rename and calculate all required columns
        summary_df.columns = [
            "Account Coding",
            "Expense Authority",
            "Total Spend (CAD)",
            "Subscriptions",
            "Projects",
        ]

        # Calculate taxes and fees (using Decimal for precise calculations)
        summary_df["Total Spend (CAD)"] = summary_df["Total Spend (CAD)"].apply(
            lambda x: Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        )

        summary_df["Vendor PST"] = summary_df["Total Spend (CAD)"] * Decimal("0.07")
        summary_df["Vendor Sub-total"] = (
            summary_df["Total Spend (CAD)"] + summary_df["Vendor PST"]
        )

        # Calculate brokerage fee
        summary_df["Brokerage Fee (6%)"] = summary_df["Total Spend (CAD)"] * Decimal(
            "0.06"
        )

        # Calculate grand total
        summary_df["Grand Total"] = (
            summary_df["Vendor Sub-total"] + summary_df["Brokerage Fee (6%)"]
        )

        # Round all decimal columns to 2 decimal places
        decimal_columns = [
            "Total Spend (CAD)",
            "Vendor PST",
            "Vendor Sub-total",
            "Brokerage Fee (6%)",
            "Grand Total",
        ]
        for col in decimal_columns:
            summary_df[col] = summary_df[col].apply(
                lambda x: x.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            )

        # Reorder columns as requested
        summary_df = summary_df[
            [
                "Account Coding",
                "Projects",
                "Total Spend (CAD)",
                "Vendor PST",
                "Vendor Sub-total",
                "Brokerage Fee (6%)",
                "Grand Total",
                "Expense Authority",
            ]
        ]

        return df, summary_df

    except Exception as e:
        print("\nDetailed error information:")
        print(f"Error type: {type(e).__name__}")
        print(f"Error message: {str(e)}")
        print(f"Error class: {e.__class__.__name__}")
        if hasattr(e, "__dict__"):
            print(f"Error attributes: {e.__dict__}")
        raise


if __name__ == "__main__":
    import argparse

    # Set up argument parser
    parser = argparse.ArgumentParser(
        description="Query Azure Cost Management for subscription costs"
    )
    parser.add_argument(
        "-m",
        "--months",
        type=int,
        default=1,
        help="Number of previous months to include (default: 1)",
    )
    parser.add_argument(
        "--mgmt-group-live",
        type=str,
        default="bcgov-managed-lz-live-landing-zones",
        help="Name of the live landing zones management group (default: bcgov-managed-lz-live-landing-zones)",
    )
    parser.add_argument(
        "--mgmt-group-decom",
        type=str,
        default="bcgov-managed-lz-live-decommissioned",
        help="Name of the decommissioned management group (default: bcgov-managed-lz-decommissioned)",
    )
    parser.add_argument(
        "--include-decom",
        action="store_true",
        help="Include the decommissioned management group in the cost recovery report",
    )
    parser.add_argument(
        "--granularity",
        type=str,
        default="Monthly",
        choices=["Daily", "Monthly", "None"],
        help="Granularity of the cost report (default: Monthly)",
    )
    parser.add_argument(
        "--output-prefix",
        type=str,
        default="azure_cost_recovery",
        help="Prefix for output files (default: azure_cost_recovery)",
    )
    parser.add_argument(
        "--start-date",
        type=str,
        default=None,
        help="Custom start date for the report period (YYYY-MM-DD). Overrides --months if set.",
    )
    parser.add_argument(
        "--end-date",
        type=str,
        default=None,
        help="Custom end date for the report period (YYYY-MM-DD). Overrides --months if set.",
    )
    args = parser.parse_args()

    mgmt_group_live = args.mgmt_group_live
    mgmt_group_decom = args.mgmt_group_decom
    include_decom = args.include_decom
    granularity = args.granularity
    output_prefix = args.output_prefix

    # Get date range based on CLI args
    if args.start_date and args.end_date:
        start = args.start_date
        end = args.end_date
    else:
        today = datetime.now()
        first_of_current = today.replace(day=1)
        last_of_previous = first_of_current - relativedelta(days=1)
        first_of_range = first_of_current - relativedelta(months=args.months)
        start = first_of_range.strftime("%Y-%m-%d")
        end = last_of_previous.strftime("%Y-%m-%d")

    print(f"\nQuerying costs for period: {start} to {end}")
    print(f"Live management group: {mgmt_group_live}")
    if include_decom:
        print(f"Decommissioned management group: {mgmt_group_decom}")
    print(f"Granularity: {granularity}")

    try:
        # Query live management group
        df_live, _ = get_subscription_costs(mgmt_group_live, start, end, granularity)
        dfs = [df_live]
        if include_decom:
            df_decom, _ = get_subscription_costs(mgmt_group_decom, start, end, granularity)
            dfs.append(df_decom)
        # Combine the results
        df = pd.concat(dfs, ignore_index=True)

        # Create summary by account coding with tax calculations (repeat logic from get_subscription_costs)
        summary_df = (
            df.groupby(["AccountCoding", "ExpenseAuthority"])
            .agg(
                {
                    "Cost": lambda x: round(sum(x), 2),
                    "SubscriptionName": lambda x: ", ".join(sorted(set(x))),
                    "LicensePlate": lambda x: ", ".join(sorted(set(x))),
                }
            )
            .reset_index()
        )

        summary_df.columns = [
            "Account Coding",
            "Expense Authority",
            "Total Spend (CAD)",
            "Subscriptions",
            "Projects",
        ]

        summary_df["Total Spend (CAD)"] = summary_df["Total Spend (CAD)"].apply(
            lambda x: Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        )
        summary_df["Vendor PST"] = summary_df["Total Spend (CAD)"] * Decimal("0.07")
        summary_df["Vendor Sub-total"] = summary_df["Total Spend (CAD)"] + summary_df["Vendor PST"]
        summary_df["Brokerage Fee (6%)"] = summary_df["Total Spend (CAD)"] * Decimal("0.06")
        summary_df["Grand Total"] = summary_df["Vendor Sub-total"] + summary_df["Brokerage Fee (6%)"]

        decimal_columns = [
            "Total Spend (CAD)",
            "Vendor PST",
            "Vendor Sub-total",
            "Brokerage Fee (6%)",
            "Grand Total",
        ]
        for col in decimal_columns:
            summary_df[col] = summary_df[col].apply(
                lambda x: x.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            )

        summary_df = summary_df[
            [
                "Account Coding",
                "Projects",
                "Total Spend (CAD)",
                "Vendor PST",
                "Vendor Sub-total",
                "Brokerage Fee (6%)",
                "Grand Total",
                "Expense Authority",
            ]
        ]

        if df.empty:
            print("\nNo data returned from the query")
        else:
            print("\nCosts by Account Coding:")
            print(summary_df.to_string(index=False))
            print("\nDetailed Subscription Costs:")
            print(df.to_string(index=False))

            # Export to CSV and Excel
            detail_csv = f"{output_prefix}_detail_{start}_to_{end}.csv"
            summary_csv = f"{output_prefix}_report_{start}_to_{end}.csv"
            summary_xlsx = f"{output_prefix}_report_{start}_to_{end}.xlsx"

            df.to_csv(detail_csv, index=False)
            summary_df.to_csv(summary_csv, index=False)

            # Create Excel writer with formatting
            with pd.ExcelWriter(summary_xlsx, engine="openpyxl") as writer:
                summary_df.to_excel(writer, sheet_name="Cost Recovery", index=False)

                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets["Cost Recovery"]

                # Define formats - Updated for openpyxl
                from openpyxl.styles import Border, Font, PatternFill, Side

                # Create styles
                header_font = Font(bold=True)
                header_fill = PatternFill(
                    start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
                )
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

                # Apply header formatting
                for col in range(1, len(summary_df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border

                # Money format for openpyxl
                money_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

                # Set column widths
                worksheet.column_dimensions["A"].width = 20  # Account Coding
                worksheet.column_dimensions["B"].width = 20  # Projects
                worksheet.column_dimensions["C"].width = 15  # Total Spend
                for col in range(4, 8):  # Other monetary columns (D through G)
                    worksheet.column_dimensions[chr(64 + col)].width = 15
                worksheet.column_dimensions["H"].width = 25  # Expense Authority

                # Apply money format to numeric columns
                for col in range(3, 8):  # Columns C through G
                    for row in range(2, len(summary_df) + 2):
                        cell = worksheet.cell(row=row, column=col)
                        cell.number_format = money_format

                # Freeze the header row
                worksheet.freeze_panes = "A2"

            print("\nResults exported to:")
            print(f"  - Detail: {detail_csv}")
            print(f"  - Summary CSV: {summary_csv}")
            print(f"  - Summary Excel: {summary_xlsx}")

    except Exception as e:
        print(f"\nFailed to retrieve cost data: {str(e)}")
